#encoding=utf8

import openpyxl
from datetime import datetime, date
# import datetime
from config import EXCEL_LOC, EMPLOYEE_SHEET_NAME, DATA_SHEET_NAME, DATE_YEAR, DATE_MONTH, DATE_DAY_FROM, DATE_DAY_COUNT

def Run():
    # workbook = openpyxl.Workbook()
    # read sheets
    workbook = openpyxl.load_workbook(EXCEL_LOC)
    sheet1 = workbook.get_sheet_by_name(DATA_SHEET_NAME)

    print sheet1.title
    # print sheet1.max_row
    # print sheet1.max_column
    row_count = sheet1.max_row
    col_count = sheet1.max_column
    print row_count
    print col_count
    # for each in xrange(1, 10):#sheet1.max_row+1):
    #     # print 123
    #     r_value = sheet1["F"+str(each)].value
    #     print r_value, type(r_value)
    #     if r_value == long(3022):
    #         print "correct!!!"
    #         print each
    #         print sheet1["E"+str(each)].value
    sheet2 = workbook.get_sheet_by_name(EMPLOYEE_SHEET_NAME)

    # create_each_day_data(sheet1, sheet2)
    create_all_datas(sheet1, sheet2)


def create_all_datas(sheet1, sheet2):
    total_employee_datas = {}
    for each in xrange(2, sheet2.max_row):
        id = sheet2['B' + str(each)].value
        employee_name = sheet2['C' + str(each)].value
        total_employee_datas[id] = {'name': employee_name, 'id': id, 'attendence_record': {}}
    # print total_employee_datas

    # total_employee_datas =
    # {'id_num': {'name': 'Bill', 'id': '1234',
    #             'attendence_record': {"10/01/18": {'start': '', 'end': '', 'working_hours': ''},
    #                                   "10/02/18": {'start': '', 'end': '', 'working_hours': ''},
    #                                   },
    #             }
    new_date_list = create_date_list(DATE_MONTH, DATE_YEAR, DATE_DAY_FROM, DATE_DAY_COUNT)
    total_employee_datas = add_date_to_total_datas(total_employee_datas, new_date_list)
    # print total_employee_datas
    print total_employee_datas.keys()

    print new_date_list
    # traverse all lines and fill datas to the sheet
    for each in xrange(2, sheet1.max_row):
        # print each
        try:
            id = str(int(sheet1["F" + str(each)].value))
        except Exception:
            continue
        date_value = sheet1["D" + str(each)].value
        time_value = sheet1["E" + str(each)].value
        r_value = sheet1["F" + str(each)].value
        n_date_value = date_value.strftime("%m/%d/%y")
        print
        if id not in total_employee_datas.keys():
            print "%s not in total_employee_datas" % id
            continue
        else:
            print "%s in total_employee_datas" % id
            print n_date_value

            if n_date_value in new_date_list:
                if not total_employee_datas[id]['attendence_record'][n_date_value]['start'] and not total_employee_datas[id]['attendence_record'][n_date_value]['end']:
                    total_employee_datas[id]['attendence_record'][n_date_value]['start'] = time_value
                    total_employee_datas[id]['attendence_record'][n_date_value]['end'] = time_value
                if time_value < total_employee_datas[id]['attendence_record'][n_date_value]['start']:
                    total_employee_datas[id]['attendence_record'][n_date_value]['start'] = time_value
                if time_value > total_employee_datas[id]['attendence_record'][n_date_value]['end']:
                    total_employee_datas[id]['attendence_record'][n_date_value]['end'] = time_value

    print total_employee_datas
    write_report(total_employee_datas)


def write_report(total_employee_datas):
    report_workbook = openpyxl.Workbook()
    report_sheet = report_workbook.create_sheet(index=0, title='finaly_report')
    report_sheet["A1"] = "Name"
    report_sheet["B1"] = "ID"
    report_sheet["C1"] = "Date"
    report_sheet["D1"] = "Start_time"
    report_sheet["E1"] = "End_time"
    report_sheet["F1"] = "Working"
    idx = 2
    for each_k, each_v in total_employee_datas.iteritems():
        for record_k, record_v in each_v['attendence_record'].iteritems():

            report_sheet['A' + str(idx)] = each_v['name']
            report_sheet['B' + str(idx)] = each_v['id']
            report_sheet['C' + str(idx)] = record_k
            report_sheet['D' + str(idx)] = record_v['start']
            report_sheet['E' + str(idx)] = record_v['end']
            if record_v['end'] and record_v['start']:
                report_sheet['F' + str(idx)] = datetime.combine(date.min, record_v['end']) - datetime.combine(date.min,
                                                                                                              record_v[
                                                                                                                  'start'])

            idx += 1

    report_workbook.save('employee_attendence_report.xlsx')

def add_date_to_total_datas(total_employee_datas, new_date_list):

    for key, each_employee in total_employee_datas.iteritems():
        attendence_record_dict = {}
        for each_date in new_date_list:
            attendence_record_dict[each_date] = {'start': '', 'end': '', 'working_hours': ''}
        each_employee['attendence_record'] = attendence_record_dict

    # print total_employee_datas
    return total_employee_datas


def create_date_list(month, year, d_start, d_end):
    date_list = []
    for each_day in xrange(d_start, d_end+1):
        # new_str = "%s/" + "%02d" % each_day + "/%s" %(str(month), str(each_day), str(year))
        new_str = str(month) + "/" + str("%02d" % each_day) + "/" + str(year)
        date_list.append(new_str)

    return date_list

def trans_date_list(date_list):
    new_date_list = []
    for each_date in date_list:
        each_date = each_date.replace("/", "_")
        new_date_list.append(each_date)
    return new_date_list

def create_each_day_data(sheet1, sheet2):
    # create_all_datas(sheet1, sheet2)
    # return
    employee_dict = {}
    for each in xrange(2, sheet2.max_row):
        employee_dict[sheet2['C'+str(each)].value] = sheet2['B'+str(each)].value
    print employee_dict

    new_date_list = create_date_list("10", "18", 1, 2)
    new_all_attendence_data = {}
    for each_date in new_date_list:
        new_all_data = []
        for each_employee, employee_id in employee_dict.items():
            employee_data = {'name': each_employee, 'id_num': employee_id, 'start_time': '', 'end_time': '', 'working': ''}
            print employee_data
            for each in xrange(2, sheet1.max_row):  # sheet1.max_row+1):
                # print "***********line: %s***********" % str(each)
                # print 123
                if not sheet1["F" + str(each)].value:
                    continue
                date_value = sheet1["D" + str(each)].value
                r_value = sheet1["F" + str(each)].value
                # print r_value, type(r_value)
                # print int(r_value), employee_data['id_num']

                n_date_value = date_value.strftime("%m/%d/%y")
                # n_date_value = datetime.strptime(date_value, "%Y-%m-%d %H:%M:%S").strftime("%m/%d/%y")
                # print date_value, n_date_value, type(date_value), each_date
                if int(r_value) == int(employee_data['id_num']):
                    print "***********line: %s***********" % str(each)
                    print date_value, n_date_value, type(date_value), type(n_date_value), each_date
                    if str(n_date_value) == str(each_date):
                        print "correct!!!"
                        time_value = sheet1["E" + str(each)].value
                        print sheet1["E" + str(each)].value
                        employee_data['end_time'] = time_value
                        if not employee_data['start_time']:
                            employee_data['start_time'] = time_value

            new_all_data.append(employee_data)
        print new_all_data
        new_all_attendence_data[each_date] = new_all_data
    # print new_all_attendence_data
    # return


    report_workbook = openpyxl.Workbook()
    # print(report_workbook.get_sheet_names())
    # sheet = report_workbook.active

    # print(sheet.title)
    #
    # sheet.title = 'report'
    # print(report_workbook.get_sheet_names())

    new_date_list = trans_date_list(new_date_list)
    for idx, each_date in enumerate(new_date_list):
        report_sheet = report_workbook.create_sheet(index=idx, title='report_' + each_date)
        print(report_workbook.get_sheet_names())
        each_date = each_date.replace("_", "/")
        attendence_data =new_all_attendence_data[each_date]

        report_sheet["A1"] = "Name"
        report_sheet["B1"] = "ID"
        report_sheet["C1"] = "Start_time"
        report_sheet["D1"] = "End_time"
        report_sheet["E1"] = "Working"


        for idx, each_person in enumerate(attendence_data):
            report_sheet['A' + str(idx + 2)] = each_person['name']
            report_sheet['B' + str(idx + 2)] = each_person['id_num']
            report_sheet['C' + str(idx + 2)] = each_person['start_time']
            report_sheet['D' + str(idx + 2)] = each_person['end_time']
            print type(each_person['end_time']), each_person['end_time']
            if each_person['end_time'] and each_person['start_time']:
                report_sheet['E' + str(idx + 2)] = datetime.combine(date.min, each_person['end_time']) - datetime.combine(date.min, each_person['start_time'])

    report_workbook.save('excel_report.xlsx')

if __name__ == "__main__":
    Run()
